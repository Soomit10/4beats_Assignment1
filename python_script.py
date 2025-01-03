import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from datetime import datetime
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# Check if the file exists and is a valid Excel file
def validate_excel_file(file_path):
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"The file {file_path} does not exist.")
    if not file_path.endswith(".xlsx"):
        raise ValueError("Please provide a valid .xlsx Excel file.")

# Set up the WebDriver for Chrome
def setup_driver():
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.get("https://www.google.com")
    return driver

# Function to get today's sheet and keywords
def get_excel_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    today = datetime.today().strftime('%A')  # Get today's day name (e.g., 'Monday')
    if today in workbook.sheetnames:
        sheet = workbook[today]
        keywords = [
            sheet.cell(row=i, column=1).value
            for i in range(2, sheet.max_row + 1)
            if sheet.cell(row=i, column=1).value is not None  # Skip empty cells
        ]
        return keywords, sheet, workbook
    else:
        raise Exception(f"No sheet found for {today} in the Excel file.")

# Function to search for a keyword and get the longest and shortest autocomplete suggestions
def search_keyword(driver, keyword):
    search_box = driver.find_element(By.NAME, "q")
    search_box.clear()
    search_box.send_keys(keyword)
    
    # Wait for suggestions to load
    suggestions = driver.find_elements(By.CSS_SELECTOR, "ul[role='listbox'] li span")
    
    # Collect autocomplete suggestions
    suggestions_list = [suggestion.text for suggestion in suggestions]

    # Get the longest and shortest suggestions
    longest = max(suggestions_list, key=len) if suggestions_list else None
    shortest = min(suggestions_list, key=len) if suggestions_list else None

    # Print statements for debugging
    print(f"Searching for: {keyword}")
    print(f"Suggestions: {suggestions_list}")
    print(f"Longest: {longest}")
    print(f"Shortest: {shortest}")

    return longest, shortest

# Write the results back to the Excel file
def write_results(sheet, workbook, file_path, results):
    for i, (keyword, longest, shortest) in enumerate(results, start=2):
        sheet.cell(row=i, column=2, value=longest)  # Write Longest Suggestion
        sheet.cell(row=i, column=3, value=shortest)  # Write Shortest Suggestion

    workbook.save(file_path)
    print("Results saved to the Excel file!")

# Main function
def main():
    file_path = "/home/soomit/Downloads/Admission Matirials/Job CV/4_beats/sample_keywords.xlsx"  # Update with the correct path to your file
    validate_excel_file(file_path)
    
    # Get keywords from today's sheet
    keywords, sheet, workbook = get_excel_data(file_path)

    if not keywords:
        print("No keywords found for today's sheet.")
    
    # Ask the user for a specific keyword to search
    user_keyword = input("Enter a keyword you want to search: ").strip()
    if user_keyword:
        keywords.append(user_keyword)  # Add the user keyword to the list of keywords

    # Set up WebDriver
    driver = setup_driver()

    results = []
    for keyword in keywords:
        # Search for the keyword and get longest and shortest suggestions
        longest, shortest = search_keyword(driver, keyword)
        results.append((keyword, longest, shortest))

        # Print keyword and its autocomplete suggestions
        print(f"Keyword: {keyword}")
        print(f"Longest Suggestion: {longest}")
        print(f"Shortest Suggestion: {shortest}")

    # Write the results back to the Excel file
    write_results(sheet, workbook, file_path, results)

    # Quit the WebDriver
    driver.quit()

if __name__ == "__main__":
    main()

